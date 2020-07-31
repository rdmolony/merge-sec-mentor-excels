from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Union, Tuple

import numpy as np
from openpyxl import load_workbook, Workbook
import pandas as pd
import prefect
from prefect import Task


class SaveDataFrameToExcelSheet(prefect.Task):
    """Saves a pandas DataFrame to a specific Excel spreadsheet

    Parameters
    ----------
    Task : prefect.Task
    """

    def run(
        self,
        excel_sheet: pd.DataFrame,
        filepath: Path,
        sheet_name: str,
        startrow: int = 1,
    ) -> pd.DataFrame:
        # Subtract 1 as Pandas is indexed at zero & Excel is indexed at one
        startrow = startrow - 1

        # https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
        book = load_workbook(filepath)
        writer = pd.ExcelWriter(filepath, engine="openpyxl")
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        excel_sheet.to_excel(
            writer, sheet_name=sheet_name, startrow=startrow, index=False,
        )
        writer.save()


class SaveDataFramesToExcel(prefect.Task):
    """A Prefect Task that saves a dictionary of pandas DataFrames to 
    corresponding sheets in an Excel Workbook

    Parameters
    ----------
    Task : prefect.Task
    """

    def run(
        self,
        dfs: Tuple[pd.DataFrame],
        filepath: Path,
        sheet_names: Tuple[str],
        header_rows: Tuple[int],
    ) -> None:
        """Saves each DataFrame in dfs to an Excel Workbook sheet corresponding 
        to the Dict key.  

        Parameters
        ----------
        dfs : Dict[str, pd.DataFrame]
        filepath : Path
        startrow : Dict[str, int], optional
            The Upper left cell row to dump each sheet's DataFrame, by default None
        """

        book = load_workbook(filepath)
        writer = pd.ExcelWriter(filepath, engine="openpyxl")
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        for df, sheet_name, header_row in zip(dfs, sheet_names, header_rows):

            df.to_excel(
                writer, sheet_name=sheet_name, startrow=header_row, index=False,
            )

        writer.save()
