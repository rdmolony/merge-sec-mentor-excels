from collections import defaultdict
from datetime import datetime
from pathlib import Path
import re
from time import strptime
from typing import List, Union, Dict
from copy import copy

from icontract import require, ensure
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import pandas as pd
from prefect import task


@task
def _get_excel_filepaths(directory) -> List[Path]:

    _all_mentor_dir_excel_files = directory.rglob("*.xlsx")
    return list(
        set(
            [
                filepath
                for filepath in _all_mentor_dir_excel_files
                for la in MENTOR_LAS
                if la in filepath.stem
            ]
        )
    )  # Using list(set()) as a hacky way of removing duplicate filepaths...


@task
def _get_excel_filepath(directory: Path, file_number) -> List[Path]:

    _all_mentor_dir_excel_files = directory.rglob("*.xlsx")
    return list(_all_mentor_dir_excel_files)[file_number]


@task
def _replace_question_marks_with_nan(df: pd.DataFrame) -> pd.DataFrame:

    return df.replace({"?": np.nan})


@task
def _drop_empty_rows(df: pd.DataFrame) -> pd.DataFrame:

    return df.dropna(how="all")


@task
def _concatenate_data_from_multiple_sheets(excel_dfs: List[pd.DataFrame]) -> pd.DataFrame:

    return pd.concat(excel_dfs, ignore_index=True)


def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=0,
    truncate_sheet=False,
    **to_excel_kwargs,
):
    """Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    Source:  https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    """

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl")

    try:
        writer.book = load_workbook(filename)

    except FileNotFoundError:
        writer.book = Workbook()

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
