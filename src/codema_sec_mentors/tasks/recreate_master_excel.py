from collections import defaultdict
from pathlib import Path
import re
from time import strptime
from typing import List, Union, Dict
from copy import copy
from shutil import copyfile

from icontract import require, ensure
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import pandas as pd
from prefect import task


@task
def _create_master_excel_from_template(template: Path, master: Path) -> None:

    copyfile(template, master)


@task
def find_header_row_and_load_sheet_to_pandas(
    filepath: Path, sheet_name: str, cell_name_in_header_row: str
) -> pd.DataFrame:

    sheet = load_workbook(filepath)[sheet_name]

    # Subtract 1 as Pandas is indexed at zero & Excel is indexed at one
    header_row = _find_cell(sheet, cell_name_in_header_row).row - 1

    sec_activity_by_month = pd.read_excel(
        filepath, sheet_name=sheet_name, header=header_row, engine="openpyxl",
    )

    sec_activity_by_month.attrs = {
        "local_authority": _extract_local_authority_name_from_filepath(filepath)
    }

    return sec_activity_by_month


def _find_cell(sheet: Worksheet, name: str) -> Cell:

    return [cell for column in sheet.columns for cell in column if cell.value == name][
        0
    ]


def _extract_local_authority_name_from_filepath(filepath: Path) -> str:

    regex = re.compile(r"SEC - CM - (\w+)")
    return regex.findall(filepath.stem)[0]


@task
def _save_to_master_excel_sheet(
    merged_df: pd.DataFrame, filepath: Path, sheet_name: str, startrow: int,
) -> pd.DataFrame:

    # Subtract 1 as Pandas is indexed at zero & Excel is indexed at one
    startrow = startrow - 1

    # https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    book = load_workbook(filepath)
    writer = pd.ExcelWriter(filepath, engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    merged_df.to_excel(
        writer, sheet_name=sheet_name, startrow=startrow, index=False,
    )
    writer.save()


@task
def _extract_summary_columns(merged_df: pd.DataFrame) -> pd.DataFrame:

    return merged_df[
        [
            "SEC Name",
            "L-P-D",
            "Maximum allocation of CM days this year",
            "CM days completed to date this year",
            "% of allocation reached",
        ]
    ]

