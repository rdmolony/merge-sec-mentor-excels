from collections import defaultdict
from datetime import datetime
from pathlib import Path
import re
from time import strptime
from typing import List, Union, Dict
from copy import copy

from icontract import require, ensure
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import pandas as pd
from prefect import Flow, task, unmapped, Parameter
from pipeop import pipes

from _filepaths import DATA_DIR, MENTOR_DIR
from _globals import MENTOR_LAS, MENTORS_BY_LA
from utilities.flow import run_flow


# Tasks
# *****

# Utility Functions
# ^^^^^^^^^^^^^^^^^


@task
def _get_excel_filepaths(directory) -> List[Path]:

    _all_mentor_dir_excel_files = directory.rglob("*.xlsx")
    return list(
        set(
            [
                filepath
                for filepath in _all_mentor_dir_excel_files
                for la in MENTOR_LAS
                if la in filepath.stem
            ]
        )
    )  # Using list(set()) as a hacky way of removing duplicate filepaths...


@task
def _get_excel_filepath(directory: Path, file_number) -> List[Path]:

    _all_mentor_dir_excel_files = directory.rglob("*.xlsx")
    return list(_all_mentor_dir_excel_files)[file_number]


@task
def _replace_question_marks_with_nan(df: pd.DataFrame) -> pd.DataFrame:

    return df.replace({"?": np.nan})


@task
def _drop_empty_rows(df: pd.DataFrame) -> pd.DataFrame:

    return df.dropna(how="all")


@task
def _concatenate_data_from_multiple_sheets(excel_dfs: List[pd.DataFrame]) -> pd.DataFrame:

    return pd.concat(excel_dfs, ignore_index=True)


def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=0,
    truncate_sheet=False,
    **to_excel_kwargs,
):
    """Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    Source:  https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    """

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl")

    try:
        writer.book = load_workbook(filename)

    except FileNotFoundError:
        writer.book = Workbook()

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


# Hourly Overview
# ---------------


@task
def _load_sec_activity_by_month_sheet_inferring_headers(filepath: Path) -> pd.DataFrame:

    sheet_name = "SEC activity by month"

    # Find row containing column headers...
    header_row_not_found = True
    row_number = 1
    sheet = load_workbook(filepath)[sheet_name]

    while header_row_not_found:

        row = sheet[row_number]
        row_values = [cell.value for cell in row[:10]]
        # search first 10 columns...
        if "SEC Name" in row_values:
            header_row_not_found = False
            break

        # If not found in first 20 rows...
        if row_number == 20:
            raise ValueError(f"No header row found in {filepath}")

        row_number += 1

    # Load excel sheet using headers found above...
    return pd.read_excel(
        filepath, sheet_name=sheet_name, header=row_number - 1, engine="openpyxl"
    )


@task
def _drop_empty_rows(excel_df: pd.DataFrame,) -> pd.DataFrame:
    """Drops all empty rows in the second column - which is column 1 before 
    headers are set...  

    Parameters
    ----------
    excel_df : pd.DataFrame

    Returns
    -------
    pd.DataFrame
    """

    return excel_df[excel_df.iloc[:, 1].notna()]


@task
def _get_rlpdt_totals(excel_df: pd.DataFrame,) -> pd.DataFrame:

    excel_df = excel_df.set_index("SEC Name")

    rlpdt_totals = ["Recruit", "Learn", "Plan", "Do", "Total"]

    return excel_df[rlpdt_totals]


# SEC hours by mentor
# -------------------


def _convert_month_string_to_number(month: str) -> int:

    return strptime(month, "%B").tm_mon


def _find_cells_corresponding_to_month(sheet: Worksheet, month_number: int) -> Cell:

    all_month_cells = [
        cell
        for column in sheet.columns
        for cell in column
        if isinstance(cell.value, datetime)
    ]

    return [cell for cell in all_month_cells if cell.value.month == month_number][0]


def _find_sec_name_cell(sheet: Worksheet) -> Cell:

    return [
        cell for column in sheet.columns for cell in column if cell.value == "SEC Name"
    ][0]


def _extract_local_authority_name_from_filepath(filepath: Path) -> str:

    regex = re.compile(r"SEC - CM - (\w+)")
    return regex.findall(filepath.stem)[0]


@task
def _load_monthly_hours_from_sec_activity_by_month_sheet(
    filepath: Path, month: str
) -> pd.DataFrame:

    sheet_name = "SEC activity by month"
    sheet = load_workbook(filepath)[sheet_name]

    month_number = _convert_month_string_to_number(month)
    month_cell = _find_cells_corresponding_to_month(sheet, month_number)

    header_row = month_cell.row + 1
    monthly_mentor_col = month_cell.column - 2
    sec_name_col = _find_sec_name_cell(sheet).column - 1

    usecols = [sec_name_col] + [*range(monthly_mentor_col, monthly_mentor_col + 10)]

    monthly_hours = pd.read_excel(
        filepath,
        sheet_name=sheet_name,
        header=header_row,
        usecols=usecols,
        engine="openpyxl",
    )

    monthly_hours.attrs = {
        "local_authority": _extract_local_authority_name_from_filepath(filepath)
    }

    return monthly_hours


@task
def _replace_empty_mentors_with_local_authority(
    monthly_hours: pd.DataFrame,
) -> pd.DataFrame:

    local_authority = monthly_hours.attrs["local_authority"]
    mentors = [col for col in monthly_hours.columns if "Mentor" in col]
    monthly_hours.loc[:, mentors] = monthly_hours[mentors].fillna(local_authority)

    return monthly_hours


@task
def _replace_empty_numeric_cells_with_zeros(
    monthly_hours: pd.DataFrame,
) -> pd.DataFrame:

    numeric_columns = monthly_hours.select_dtypes("number").columns.tolist()
    monthly_hours.loc[:, numeric_columns] = monthly_hours.fillna(0)

    return monthly_hours


@task
def _label_monthly_hours_as_planned_and_acheived(
    monthly_hours: pd.DataFrame,
) -> pd.DataFrame:

    monthly_hours = monthly_hours.set_index("SEC Name")

    monthly_hours.columns = pd.MultiIndex.from_product(
        [["Planned", "Achieved"], ["County Mentor", "Recruit", "Learn", "Plan", "Do"]],
        names=["Status", ""],
    )

    return monthly_hours.stack(0)


@task
def _calculate_total_monthly_hours(monthly_hours: pd.DataFrame,) -> pd.DataFrame:

    return monthly_hours.assign(
        Total=monthly_hours[["Recruit", "Learn", "Plan", "Do"]].sum(axis=1)
    )


@task
def _get_total_monthly_hours_by_mentor(monthly_hours: pd.DataFrame,) -> pd.DataFrame:

    return monthly_hours.pivot_table(values="Total", index=["County Mentor", "Status"])


# Recreate the master excel
# -------------------------

@task
def _copy_guidance_sheet(template: Path, todays_master: Path):

    template_workbook = load_workbook(template)

    try:
        todays_workbook = load_workbook(todays_master)

    except FileNotFoundError:
        todays_workbook = Workbook()

    guidance_sheet = todays_workbook.create_sheet("Guidance")

    for row in template_workbook["Guidance"]:
        for cell in row:
            guidance_sheet[cell.coordinate] = cell.value

        if cell.has_style:
            guidance_sheet[cell.coordinate].style = cell.style
            # guidance_sheet[cell.coordinate].font = copy(cell.font)
            # guidance_sheet[cell.coordinate].border = copy(cell.border)
            # guidance_sheet[cell.coordinate].fill = copy(cell.fill)
            # guidance_sheet[cell.coordinate].number_format = copy(cell.number_format)
            # guidance_sheet[cell.coordinate].protection = copy(cell.protection)
            # guidance_sheet[cell.coordinate].alignment = copy(cell.alignment)

         

    todays_workbook.save(todays_master)


@task
def _load_sec_activity_by_month_sheets_inferring_headers(filepath: Path) -> pd.DataFrame:

    sheet_name = "SEC activity by month"
    sheet = load_workbook(filepath)[sheet_name]

    # -1 as Pandas is indexed at zero & Excel is indexed at one
    header_row = _find_sec_name_cell(sheet).row - 1

    sec_activity_by_month = pd.read_excel(
        filepath, sheet_name=sheet_name, header=header_row, engine="openpyxl",
    )

    sec_activity_by_month.attrs = {
        "local_authority": _extract_local_authority_name_from_filepath(filepath)
    }

    return sec_activity_by_month


@task
def _save_sec_activity_by_month_to_excel(merged_df: pd.DataFrame, filepath: Path) -> pd.DataFrame:

    append_df_to_excel(filepath, merged_df, sheet_name="SEC activity by month", startrow=7)


# ETL Flow
# ********

RESULTS_DIR = DATA_DIR / "results"
TODAYS_MASTER_EXCEL = DATA_DIR / "results" / f"master-{datetime.today().strftime('%d-%m-%y')}.xlsx"
TEMPLATE_MASTER_EXCEL = DATA_DIR / "master_template.xlsx"  

@pipes
def recreate_master_excel_flow() -> Flow:

    with Flow("Recreate Master Excel") as flow:

        filepaths = _get_excel_filepaths(MENTOR_DIR)
        filepath = _get_excel_filepath(MENTOR_DIR, 1)

        _copy_guidance_sheet(TEMPLATE_MASTER_EXCEL, TODAYS_MASTER_EXCEL)

        sec_by_month_sheet_data = (
            _load_sec_activity_by_month_sheets_inferring_headers.map(filepaths)
            >> _replace_question_marks_with_nan.map
            >> _drop_empty_rows.map
            >> _concatenate_data_from_multiple_sheets
        )
        _save_sec_activity_by_month_to_excel(sec_by_month_sheet_data, TODAYS_MASTER_EXCEL)
        

    return flow


@pipes
def monthly_report_flow() -> Flow:

    with Flow("Compile monthly SEC mentor report") as flow:

        filepaths = _get_excel_filepaths(MENTOR_DIR)

        file_number = Parameter("file_number")
        filepath = _get_excel_filepath(MENTOR_DIR, file_number)

        # hourly_overview = (
        #     _load_sec_activity_by_month_sheet_inferring_headers.map(filepaths)
        #     >> _drop_empty_rows.map
        #     >> _get_rlpdt_totals.map
        #     >> _concatenate_dataframes
        # )
        # _save_dataframe_to_excel(hourly_overview, RESULTS_DIR, "total_sec_hours.xlsx")

        month = Parameter("month")
        monthly_hours = (
            _load_monthly_hours_from_sec_activity_by_month_sheet.map(
                filepaths, unmapped(month)
            )
            >> _replace_question_mark_with_nan.map()
            >> _drop_empty_rows.map
            >> _replace_empty_mentors_with_local_authority.map
            >> _replace_empty_numeric_cells_with_zeros.map
            >> _label_monthly_hours_as_planned_and_acheived.map
            >> _calculate_total_monthly_hours.map
            >> _concatenate_dataframes
        )
        _save_dataframe_to_excel(monthly_hours, RESULTS_DIR, "monthly_hours.xlsx")

        monthly_hours_by_mentor = _get_total_monthly_hours_by_mentor(monthly_hours)
        _save_dataframe_to_excel(
            monthly_hours_by_mentor, RESULTS_DIR, "monthly_hours_by_mentor.xlsx"
        )

    return flow
